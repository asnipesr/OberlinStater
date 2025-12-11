##TO DO FEATURE PLAN

# Calulcate Advanced Metrics -- create a similar one for regular statistics
# make it so that it can append on to existing file
# give rest of players 0 if not clicked
# make into excel file that looks similar to one existing
# Auto Save
# Live Feed
# Aggregate Stats Generator
# import csv and work from there in the case of a crash
# https://stackoverflow.com/questions/10020885/creating-a-popup-message-box-with-an-entry-field

import pygame
import pygame.locals 
from button import button, text
from datetime import date
import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from pathlib import Path 
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import os

ROOT = tk.Tk()
ROOT.withdraw()
CONTINUE_WB = False

# Gets wanted filename
today = date.today()
USER_INP = simpledialog.askstring(title="User Input", prompt="Name of File (don't include extension)")
if not USER_INP or USER_INP.strip() == "":
    base_name = f"./{today.month}-{today.day}statsheet"
else:
    base_name = USER_INP.strip()
    
file_name = base_name + ".xlsx"

# Checks if filename exists
# OPTIONS:
# Continue Working
# Add new worksheet
# Overwrite existing
if Path(file_name).is_file():
    ADDSHEET_INP = messagebox.askyesno(title="Pick Either", message="This file already exists. Do you want to add a new sheet to the workbook")

    if not ADDSHEET_INP:
        APPEND_INP = messagebox.askyesno(title="Continue Working", message="Do you want to continue working on an existing workbook?")
        
        if not APPEND_INP:
            OVERWRITE_INP = messagebox.askyesno(title="Pick one", message="Are you sure? Selecting YES will overwrite existing files! Selecting NO will create a copy")
            if not OVERWRITE_INP:
                count = 0
                if not OVERWRITE_INP:   
                    while Path(file_name).is_file():
                        count+=1
                        file_name = f"./{base_name}_{count}.xlsx"
    else:
        CONTINUE_WB = True
        wb = load_workbook(filename=file_name)
        SHEETNAME_INP = simpledialog.askstring(title="New Sheet Input", prompt="Enter name of name of the new sheet (can't use '/' character)")
        
        if not SHEETNAME_INP or SHEETNAME_INP.strip() == "":
            sheet_name = f"./{today.month}-{today.day}statsheet"
        else:
            sheet_name = SHEETNAME_INP.strip()
        
        # same thing with checking to append or overwrite sheet
        
        count = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name} {count}"
            count+=1  

# check it out
print("Saving name", file_name)
print("\n")

messagebox.showinfo(title="Instructions", message="Stats won't save on every input anymore. Stats will only save at the end once quitting the statting application by hitting the red X in the top left corner.\nUse CMD+S to save at any time\nUse CMD+Z to undo last stat entry\nClick on a stat button, then click on a player button to record that stat for that player")

#initializing pygame
pygame.init()

#setting screen width and height
SCREEN_WIDTH = 800
SCREEN_HEIGHT = 800
ROSTER_SIZE = 20

#creating game window
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("Stat Tracker")

#creating text for title
screen.fill((0,0,0))
pygame.font.init()
textfont = pygame.font.SysFont("monospace", 50)

textTBR = textfont.render("STAT TRACKER", 1, (255,255,255))
screen.blit(textTBR, (220, 10))

# create a surface object, image is drawn on it.
imp = pygame.image.load("goyeo.png").convert()
imp = pygame.transform.scale(imp,(200,200))
 
# Using blit to copy content from one surface to other
screen.blit(imp, (290, 100))
 
# paint screen one time
pygame.display.flip()

run = True
selected = False
global curr
   
# Sends stats to file and formats worksheet
def send_to_file(stats, wb=None, sheet_name=""):
    header = ["PLAYER","MADE THREES", "ATTEMPTS THREES", "PCT\nTHREES", "MADE MIDDY", "ATTEMPTS MIDDY", "PCT\nMIDDY", "MADE LAYUP", "ATTEMPTS LAYUP", "PCT\nLAYUP", "MADE\nFTs", "ATTEMPTS FTs", "PCT\nFTs", "AST", "OBIE AST", "TO", "OREB", "DREB", "REB", "STL", "BLK\nW-UP", "DEFL", "BLOW BY\nMADE IN FACE", "BAD ROTATION", "DRAW FL", "FOUL", "TEAM WIN"]
    # multipliers = [3,-1,2,-1,1,-1,1,2,2,-3,1,2,1,0,2,2,1,3,1,-1,-1,1]
    
    # Create workbook if none exists (doesn't want to append)
    if wb is None:
        wb = Workbook()
        ws = wb.active
        if ws.title == "Sheet":
            wb.remove(ws)
    
    # Clear cells in worksheet or creates new one
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None
    else:
        ws = wb.create_sheet(title=sheet_name)
        
    NUMBER_OF_STATS = len(header) - 1
    EXTRA_ROWS = 1 # PLAYER NAME
    col_letter = get_column_letter(NUMBER_OF_STATS + EXTRA_ROWS) 
    
    # Generate Top Header
    ws.merge_cells(f"A1:{col_letter}1")
    ws['A1'] = "Oberlin College Basketball"
    ws['A1'].font = Font(name=" Oswald", size=22, bold=True, color=("FFFFFF"))
    ws['A1'].fill = PatternFill(start_color="81192e", end_color="81192e",fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Generate secondary header
    ws.merge_cells(f"A2:{col_letter}2")
    ws['A2'] = f"Oberlin Practice Stats - {today.month}/{today.day}"
    ws['A2'].font = Font(name="Oswald", size=16, bold=False, italic=True, color="000000")
    ws['A2'].fill = PatternFill(start_color="ffb800", end_color="ffb800",fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    # Formats Header Row 
    start_header_row = 3
    ws.row_dimensions[3].height = 38.00
    for col_index, value in enumerate(header, start=1):
        cell = ws.cell(row=start_header_row, column=col_index, value=value)
        cell.font = Font(name="Oswald", bold=True, italic=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="000000"))

        
    # Formats Bottom Row
    ws.merge_cells(f"A{ROSTER_SIZE+4}:{col_letter}{ROSTER_SIZE+4}")
    ws[f'A{4+ROSTER_SIZE}'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.row_dimensions[4 + ROSTER_SIZE].height = 26.00
    
    # ----- PLAYER ROWS -----
    # Generates rows of player statistic data
    player_rows = []
    for person in stats.keys():
        values = stats[person] 
        row_data = [person] + values
        player_rows.append(row_data)
        
    # Generates row for player if they are not already on the board
    for player in players:
        name = player["name"]
        if name not in stats:
            zero_stats = [0] * NUMBER_OF_STATS
            row_data = [name] + zero_stats
            player_rows.append(row_data)
    
    # Sorts the players by first name 
    player_rows.sort(key=lambda row:row[0])
    start_row = 4
    for r_index, row in enumerate(player_rows, start=start_row):
        for c_index, value in enumerate(row, start=1):
            ws.cell(row=r_index, column=c_index, value=value)   
        
    # Styles and Formats Statistics Cells
    # Gives different shades for cells in every other row
    for r in range(4, 4+ROSTER_SIZE):
        ws[f"A{r}"].font = Font(name="Oswald", size=12)
        ws[f"A{r}"].alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22.00
        if r % 2 == 1:
            ws[f"A{r}"].fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
        for c in range(2, len(header)+1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Oswald", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(right=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000")) 
            if r % 2 == 1:
                cell.fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
        ws[f"A{r}"].border = Border(right=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"))

    # Styles and Colors Percentage Rows
    for r in range(4, 4+ROSTER_SIZE):
        for c in range(2, len(header)+1):
            col_letter = get_column_letter(c)
            cell = ws.cell(row=r, column=c)
            
            # --- THREE POINT PERCENTAGE ---
            if col_letter == 'D':
                if cell.value > .35:
                    cell.font = Font(color="00C409")
                elif cell.value <= .35 and cell.value > .3:
                    cell.font = Font(color="E6A61B")
                else:
                    cell.font = Font(color="FF0000")
                    
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
                cell.border = Border(right=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"))
            
            # --- MID RANGE PERCENTAGE ---
            elif col_letter == 'G':
                if cell.value > .45:
                    cell.font = Font(color="00C409")
                elif cell.value <= .45 and cell.value > .4:
                    cell.font = Font(color="E6A61B")
                else:
                    cell.font = Font(color="FF0000")
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
                cell.border = Border(right=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"))
                
            # --- LAYUP PERCENTAGE ---
            elif col_letter == 'J':
                if cell.value > .65:
                    cell.font = Font(color="00C409")
                elif cell.value <= .65 and cell.value > .50:
                    cell.font = Font(color="E6A61B")
                else:
                    cell.font = Font(color="FF0000")
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
                cell.border = Border(right=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"))
                
            # --- FREE THROW PERCENTAGE ---
            elif col_letter == 'M':
                if cell.value > .80:
                    cell.font = Font(color="00C409")
                elif cell.value <= .80 and cell.value > .70:
                    cell.font = Font(color="E6A61B")
                else:
                    cell.font = Font(color="FF0000")
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
                cell.border = Border(right=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"))
                
            elif col_letter == 'P' or col_letter == 'S':
                cell.border = Border(right=Side(style="thick", color="000000"), bottom=Side(style="thin", color="000000"))
                    
            
            
    # Spaces cells width and height correctly  
    OFFSET = 0.83
    for col in ws.columns:
        column = get_column_letter(col[0].column)  # Convert 1 -> 'A', etc.
        if column == 'A':
            ws.column_dimensions[column].width = 16.00 + OFFSET 
        elif column == 'S':
            ws.column_dimensions[column].width = 5.00 + OFFSET
        elif column == 'W':
            ws.column_dimensions[column].width = 11.00 + OFFSET
        elif column > 'A' and column <= 'M' or column == "X" or column == "U": 
            ws.column_dimensions[column].width = 8.00 + OFFSET   
        elif column == 'AA':
            ws.column_dimensions[column].width = 7.00 + OFFSET  
        else:
            ws.column_dimensions[column].width = 7.00 + OFFSET  
    
    wb.save(file_name)
    
def find(lst, num):
    for dictionary in lst:
        if dictionary["number"] == num:
            return dictionary
    return {}

def find_option(lst, name):
    for dictionary in lst:
        if dictionary["name"] == name:
            return dictionary
    return {}

stats = {}
global stat_records
stat_records = []
def Number(num):
    global stat_records
    new_stats = {}
    for i in stats.keys():
        new_stats[i] = stats[i].copy()
    stat_records.append(new_stats)
    # print(stat_records)
    # print("Records")
    player_dict = find(players, num)
    name = player_dict["name"]
    stats_dict = find_option(options, curr)
    if name in stats:
        #print(stats_dict["index"])
        stats[name][stats_dict["index"]] += 1
        if stats_dict["index"] == 15 or stats_dict["index"] == 16:
            stats[name][17] += 1
        
        if stats_dict["index"] == 0 or stats_dict["index"] == 3 or stats_dict["index"] == 6 or stats_dict["index"] == 9:
            stats[name][stats_dict["index"] + 1] += 1
        
        if stats_dict["index"] == 0 or stats_dict["index"] == 1:
            stats[name][2] = stats[name][0] / stats[name][1]
        elif stats_dict["index"] == 3 or stats_dict["index"] == 4:
            stats[name][5] = stats[name][3] / stats[name][4]
        elif stats_dict["index"] == 6 or stats_dict["index"] == 7:
            stats[name][8] = stats[name][6] / stats[name][7]
        elif stats_dict["index"] == 9 or stats_dict["index"] == 10:
            stats[name][11] = stats[name][9] / stats[name][10]
    else:
        stats[name] = [0] * 26
        stats[name][stats_dict["index"]] = 1
        if stats_dict["index"] == 15 or stats_dict["index"] == 16:
            stats[name][17] += 1
            
        if stats_dict["index"] == 0 or stats_dict["index"] == 3 or stats_dict["index"] == 6 or stats_dict["index"] == 9:
            stats[name][stats_dict["index"] + 1] += 1
        
        if stats_dict["index"] == 0 or stats_dict["index"] == 1:
            stats[name][2] = stats[name][0] / stats[name][1]
        elif stats_dict["index"] == 3 or stats_dict["index"] == 4:
            stats[name][5] = stats[name][3] / stats[name][4]
        elif stats_dict["index"] == 6 or stats_dict["index"] == 7:
            stats[name][8] = stats[name][6] / stats[name][7]
        elif stats_dict["index"] == 9 or stats_dict["index"] == 10:
            stats[name][11] = stats[name][9] / stats[name][10]

    print(name + " -- " + stats_dict["name"])
    
    
# "WIN" ,"FGM", "FGA", "3PM", "3PA", "AST", "ORB", "DRB","STL","BLK","TOV"]
# "M3", "M NR2", "RIM2", "FTS", "AST", "OBIE AST", "TO", "PT", "OREB", "DREB", "REB", "STL", "BLK_WUP", "DEFL", "DRAW FL", "FOUL", "BLOW BY", "TEAM WIN", "POSS"
def MADE_THREE():
    global curr
    curr = "MADE THREE"
def MISS_THREE():
    global curr
    curr = "MISSED THREE"
def MADE_NR2():
    global curr
    curr = "MADE MIDDY"
def MISS_NR2():
    global curr
    curr = "MISSED MIDDY"
def MADE_R2():
    global curr
    curr = "MADE LAYUP"
def MISS_R2():
    global curr
    curr = "MISSED LAYUP"
def MADE_FTs():
    global curr
    curr = "MADE FT"
def MISS_FTs():
    global curr
    curr = "MISSED FT"
def AST():
    global curr
    curr = "AST"
def Obie_AST():
    global curr
    curr = "OBIE AST"
def TOV():
    global curr
    curr = "TOV"
def OREB():
    global curr
    curr = "OREB"
def DREB():
    global curr
    curr = "DREB"
def STL():
    global curr
    curr = "STL"
def BLK_WALLUP():
    global curr
    curr = "BLK/WALL-UP"
def DEFL():
    global curr
    curr = "DEFL"
def DRAW_FL():
    global curr
    curr = "DRAW FOUL"
def FOUL():
    global curr
    curr = "FOUL"
def BLOW_BY():
    global curr
    curr = "BLOW BY / MAKE IN FACE"
def BAD_ROTATION():
    global curr
    curr = "BAD ROTATION"
def WIN():
    global curr
    curr = "TEAM WIN"
# def POSS():
#     global curr
#     curr = "POSS"

# "M3", "M NR2", "RIM2", "FTS", "AST", "OBIE AST", "TO", "PT", "OREB", "DREB", "REB", "STL", "BLK_WUP", "DEFL", "DRAW FL", "FOUL", "BLOW BY", "TEAM WIN", "POSS"
options = [
    { "name" : "MADE THREE",
      "function": MADE_THREE,
      "index": 0
    },
    { "name" : "MISSED THREE",
      "function": MISS_THREE,
      "index": 1
    },
    { "name" : "MADE MIDDY",
      "function": MADE_NR2,
      "index": 3
    },
    { "name" : "MISSED MIDDY",
      "function": MISS_NR2,
      "index": 4
    },
    { "name" : "MADE LAYUP",
      "function": MADE_R2,
      "index": 6
    },
    { "name" : "MISSED LAYUP",
      "function": MISS_R2,
      "index": 7
    },
    { "name" : "MADE FT",
      "function": MADE_FTs,
      "index": 9
    },
    { "name" : "MISSED FT",
      "function": MISS_FTs,
      "index": 10
    },
    { "name" : "AST",
      "function": AST,
      "index": 12
    },
    { "name" : "OBIE AST",
      "function": Obie_AST,
      "index": 13
    },
    { "name" : "TOV",
      "function": TOV,
      "index": 14
    },
    { "name" : "OREB",
      "function": OREB,
      "index": 15
    },
    { "name" : "DREB",
      "function": DREB,
      "index": 16
    },
    #total rebounds is index 17
    { "name" : "STL",
      "function": STL,
      "index": 18
    },
    { "name" : "BLK/WALL-UP",
      "function": BLK_WALLUP,
      "index": 19
    },
    { "name" : "DEFL",
      "function": DEFL,
      "index": 20
    },
    { "name" : "BLOW BY / MAKE IN FACE",
      "function": BLOW_BY,
      "index": 21
    },
    { "name" : "BAD ROTATION",
      "function": BAD_ROTATION,
      "index": 22
    },
    { "name" : "DRAW FOUL",
      "function": DRAW_FL,
      "index": 23
    },
    { "name" : "FOUL",
      "function": FOUL,
      "index": 24
    },
    { "name" : "TEAM WIN",
      "function": WIN,
      "index": 25
    }
    # ,
    # { "name" : "POSS",
    #   "function": POSS,
    #   "index": 26
    # }
]

players = [
    {"number": "0",
     "nickname": "Gus",
     "function": Number,
     "name": "Gus Donahue",
     "img": "./players/0gus.jpg"
    },
    {"number": "1",
     "nickname": "ZB",
     "function": Number,
     "name": "Zach Bronson",
     "img": "./players/1zach.jpg"
     },
    {"number": "3",
     "nickname": "Alasan",
     "function": Number,
     "name": "Alasan Njie-Morgan",
     "img": "./players/3alasan.jpg",
    },
    {"number": "4",
     "nickname": "Liam",
     "function": Number,
     "name": "Liam Gray",
    "img": "./players/4liam.jpg"
    },
    {"number": "5",
     "nickname": "Matty Ice",
     "function": Number,
     "name": "Matt Andreopolous",
     "img": "./players/5matt.jpg"},
    {"number": "10",
     "nickname": "Nav",
     "function": Number,
     "name": "Adam Navarre",
     "img": "./players/10adam.jpg"},
    {"number": "12",
     "nickname": "Smitty",
     "function": Number,
     "name": "Zach Smith",
     "img": "./players/12smitty.jpg"},
    {"number": "13",
     "nickname": "Kyryl",
     "function": Number,
     "name": "Kyryl Streltsov",
     "img": "./players/13kyryl.jpg"
    },
    {"number": "14",
     "nickname": "Matt",
     "function": Number,
     "name": "Matthew Callahan",
     "img": "./players/14matt.jpg"
    },
    {"number": "15",
     "nickname": "Mo",
     "function": Number,
     "name": "Mouhamed Toure",
     "img": "./players/15mo.jpg"
    },
    {"number": "20",
     "nickname": "JJ",
     "function": Number,
     "name": "JJ Gray",
     "img": "./players/20jj.jpg"
     },
    {"number": "21",
     "nickname": "Will",
     "function": Number,
     "name": "William Manfredi",
     "img": "./players/21will.jpg"
    },
    {"number": "22",
     "nickname": "Shea",
     "function": Number,
     "name": "Shea Laursen",
     "img": "./players/22shea.jpg"
     },
    {"number": "23",
     "nickname": "Rob",
     "function": Number,
     "name": "Rob Magner",
     "img": "./players/23rob.jpg"
     },
    {"number": "24",
     "nickname": "JRen",
     "function": Number,
     "name": "Jackson Reynolds",
     "img": "./players/24jackson.jpg"
     },
    {"number": "30",
     "nickname": "Noah",
     "function": Number,
     "name": "Noah Kim",
     "img": "./players/30noah.jpg"
    },
    {"number": "33",
     "nickname": "Uros",
     "function": Number,
     "name": "Uros Petrusic",
     "img": "./players/33uros.jpg"
    },
    {"number": "34",
     "nickname": "Dut",
     "function": Number,
     "name": "Dut Lual",
     "img": "./players/34dut.jpg"
    },
    {"number": "35",
     "nickname": "Milun",
     "function": Number,
     "name": "Milun Micanovic",
     "img": "./players/35milun.jpg"
    },
    {"number": "50",
     "nickname": "Sam",
     "function": Number,
     "name": "Samuel Kamenko",
     "img": "./players/50sam.jpg"
    },
]

button_list = []
x = 80
x_increment = 125
y = 425
row = 0
y_increment = 75
count = 0

# Creates initial buttons - Statistics
for option in options:
    pos_x = x + (count * x_increment)
    pos_y = y + (row * y_increment)
    
    curr_button = button(position = (pos_x, pos_y), size=(100, 50), clr=(220, 220, 220), cngclr=(255, 0, 0), func=option["function"], text=option["name"])
    count+=1
    
    if pos_x > 700:
        count = 0
        row += 1
        
    button_list.append(curr_button)

player_list = []
player_images = []
x2 = 80
x_increment2 = 125
y2 = 265
row = 0
y_increment2 = 160
count = 0

# Create player buttons
for option in players:
    if "img" in players:
        img = pygame.image.load(option["img"]).convert_alpha()
        img = pygame.transform.scale(img, (100,125))
        option["img"] = img

for option in players:
    pos_x = x2 + (count * x_increment2)
    pos_y = y2 + (row * y_increment2)
    
    curr_button = button(position = (pos_x, pos_y), size=(100, 50), clr=(220, 220, 220), cngclr=(255, 0, 0), func=option["function"], text=option["number"])
    count+=1
    
    if pos_x > 700:
        count = 0
        row += 1
    player_list.append(curr_button)
    
    # Places image to corresponding player 
    if "img" in option:
        img = pygame.image.load(option["img"]).convert_alpha()
        img = pygame.transform.scale(img, (70,87))
        img_rect = img.get_rect(center=(pos_x, pos_y - 73))
        # player_images.append((img, img_rect))
        player_images.append({"img": img, "rect": img_rect, "player_number": option["number"], "func": option["function"]})

def save():
    if stats == {}:
        print("No Stats To Save")
    else:
        if CONTINUE_WB:
            send_to_file(stats,wb=wb,sheet_name=sheet_name)
        else:
            send_to_file(stats,sheet_name="Stat Sheet")
        print("\nSaved to ", file_name, "\nPath: ", os.getcwd())

def new_game():
    stats = {}
    players = []

while run:
    in_cycle = True
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            save()
            run = False
        elif event.type == pygame.KEYDOWN:
            if pygame.key.get_mods() & pygame.KMOD_META:
                if event.key == pygame.K_s:
                    save()
                if event.key == pygame.K_z:
                    print("Undo")
                    if len(stat_records) > 0:
                        stats = stat_records[len(stat_records)-1].copy()
                        stat_records = stat_records[0:-1]
                        print(stats)
                if event.key == pygame.K_1:
                    quit
        elif event.type == pygame.MOUSEBUTTONDOWN:
                if event.button == 1:
                    pos = pygame.mouse.get_pos()
                    pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 0, 800, 800))
                    if not selected and not press:
                        for b in button_list:
                            if b.rect.collidepoint(pos):
                                selected = True
                                press = True
                                pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 300, 800, 500))
                                pygame.display.flip()
                                b.call_back()
                    if selected and not press:
                        for b in player_list:
                            if b.rect.collidepoint(pos):
                                selected = False
                                press = True
                                pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 0, 800, 800))
                                pygame.display.flip()
                                b.call_back(b.txt)
                        for p in player_images:
                            if p["rect"].collidepoint(pos):
                                selected = False
                                press = True
                                pygame.draw.rect(screen, (0,0,0), pygame.Rect(0, 0, 800, 800))
                                pygame.display.flip()
                                p["func"](p["player_number"])
        else:
            press = False
                
    if not selected:
        if in_cycle:
            screen.blit(textTBR, (220, 10))
            screen.blit(imp, (290, 100))
        for b in button_list:
            b.draw(screen)
    if selected:
        for b in player_list:
            b.draw(screen)
        for p in player_images:
            screen.blit(p["img"], p["rect"])

    pygame.display.update()

pygame.quit()